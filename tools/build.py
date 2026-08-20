#!/usr/bin/env python3
"""Gera o index.html do app a partir da planilha do checklist.

Uso:
    python3 tools/build.py [caminho_da_planilha]

Lê a planilha (padrão: data/hotwheels_bmw_porsche_ferrari_checklist.xlsx),
converte as abas em JSON compacto e injeta no app/template.html no lugar do
marcador __HW_DATA_JSON__, gravando o resultado em index.html (arquivo único,
funciona offline e no GitHub Pages).
"""
import json
import sys
import warnings
from datetime import date
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data" / "hotwheels_bmw_porsche_ferrari_checklist.xlsx"
TOONED = ROOT / "data" / "hot_wheels_tooned_variants_2026-08-19.json"
TEMPLATE = ROOT / "app" / "template.html"
OUT = ROOT / "index.html"

# Abas de itens: (nome da aba, chave, prefixo de id)
SHEETS = [
    ("BMW", "bmw", "b"),
    ("Porsche", "porsche", "p"),
    ("Ferrari", "ferrari", "f"),
    ("Outras linhas", "outras", "o"),
]

# Ordem fixa das 24 colunas das abas de itens (linha 4 = cabeçalho)
N_COLS = 24


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def read_items(ws):
    headers = [clean(c.value) for c in ws[4]][:N_COLS]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        vals = [clean(v) for v in row[:N_COLS]]
        if all(v is None for v in vals):
            continue
        rows.append(vals)
    return headers, rows


def load_tooned():
    """Lê o seed Tooned e valida a integridade antes de embutir no app.

    O seed é fonte de verdade só do catálogo; a coleção do usuário vive
    separada (localStorage, por variant.id), então trocar este arquivo por
    uma versão nova não apaga nada do que o usuário marcou.
    """
    data = json.loads(TOONED.read_text(encoding="utf-8"))
    castings = data["castings"]
    variants = data["variants"]

    erros = []
    vistos = set()
    for v in variants:
        if v["id"] in vistos:
            erros.append(f"variant.id duplicado: {v['id']}")
        vistos.add(v["id"])
    ids_castings = set()
    for c in castings:
        if c["casting_id"] in ids_castings:
            erros.append(f"casting_id duplicado: {c['casting_id']}")
        ids_castings.add(c["casting_id"])
    for v in variants:
        if v["casting_id"] not in ids_castings:
            erros.append(f"casting_id órfão em {v['id']}: {v['casting_id']}")
    if erros:
        raise SystemExit("Seed Tooned inválido:\n  " + "\n  ".join(erros[:20]))

    checklist = [v for v in variants if v.get("primary_checklist") and v.get("status") == "released"]
    producao = [v for v in variants if v.get("variation_type") == "production_variant"]
    print(f"Tooned: {len(castings)} castings, {len(variants)} variações, "
          f"{len(checklist)} no checklist principal, {len(producao)} variações de produção, "
          f"{len(data.get('segment_name_only_2014_2015', []))} de 2014-2015, "
          f"{len(data.get('review_queue', []))} em revisão")

    # só o que o app usa, para não inflar o arquivo único
    return {
        "asOf": data["meta"]["as_of"],
        "castings": castings,
        "variants": variants,
        "segment2014": data.get("segment_name_only_2014_2015", []),
        "familias": data.get("related_style_families", []),
        "pendentes": data.get("upcoming_or_pending", []),
        "revisao": data.get("review_queue", []),
        "fontes": data.get("sources", {}),
        "totais": {"castings": len(castings), "variantes": len(variants),
                   "checklist": len(checklist), "producao": len(producao)},
    }


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    marcas = []
    headers_ref = None
    for sheet, key, prefix in SHEETS:
        ws = wb[sheet]
        headers, rows = read_items(ws)
        if headers_ref is None:
            headers_ref = headers
        elif headers != headers_ref:
            raise SystemExit(f"Cabeçalhos da aba {sheet} diferem da primeira aba: {headers}")
        marcas.append({
            "key": key,
            "prefixo": prefix,
            "nome": clean(ws["A1"].value) or sheet,
            "aba": sheet,
            "desc": clean(ws["A2"].value) or "",
            "itens": rows,
        })
        print(f"{sheet}: {len(rows)} itens")

    # Índice de castings (cabeçalho na linha 4)
    ws = wb["Castings"]
    castings = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        vals = [clean(v) for v in row[:10]]
        if all(v is None for v in vals):
            continue
        castings.append(vals)
    print(f"Castings: {len(castings)}")

    # Guia (cabeçalho na linha 3: Tema / Critério adotado)
    ws = wb["Guia"]
    guia = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        tema, crit = clean(row[0]), clean(row[1])
        if tema or crit:
            guia.append([tema, crit])

    # Fontes (cabeçalho na linha 3: Fonte / Uso / URL)
    ws = wb["Fontes"]
    fontes = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = [clean(row[0]), clean(row[1]), clean(row[2])]
        if any(vals):
            fontes.append(vals)

    painel = wb["Painel"]
    data = {
        "titulo": clean(painel["A1"].value) or "Coleção Hot Wheels",
        "nota": clean(painel["A3"].value) or "",
        "geradoEm": date.today().isoformat(),
        "cols": ["own", "pri", "ano", "casting", "modelo", "serie", "nserie", "ncol",
                 "toy", "cor", "tampo", "rodas", "base", "vidros", "interior", "pais",
                 "excl", "tipo", "rar", "sit", "notas", "escopo", "fonte", "pagina"],
        "rotulos": headers_ref,
        "marcas": marcas,
        "castings": castings,
        "guia": guia,
        "fontes": fontes,
    }

    def embutir(obj):
        # Evita fechamento prematuro da tag <script> por strings dos dados
        return json.dumps(obj, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")

    html = TEMPLATE.read_text(encoding="utf-8")
    for marker, obj in (("__HW_DATA_JSON__", data), ("__HW_TOONED_JSON__", load_tooned())):
        if marker not in html:
            raise SystemExit(f"Marcador {marker} não encontrado no template")
        html = html.replace(marker, embutir(obj), 1)
    OUT.write_text(html, encoding="utf-8")

    total = sum(len(m["itens"]) for m in marcas)
    print(f"OK: {OUT.name} gerado com {total} itens ({OUT.stat().st_size/1e6:.2f} MB)")


if __name__ == "__main__":
    main()
